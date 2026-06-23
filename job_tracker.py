import os
import json
import base64
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from google import genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
from dotenv import load_dotenv
load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
MAX_EMAILS = 2000

def load_scanned_ids():
    if os.path.exists('scanned_ids.json'):
        with open('scanned_ids.json', 'r') as f:
            return set(json.load(f))
    return set()

def save_scanned_ids(ids):
    existing = load_scanned_ids()
    all_ids = existing.union(ids)
    with open('scanned_ids.json', 'w') as f:
        json.dump(list(all_ids), f)

def authenticate_gmail():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('gmail', 'v1', credentials=creds)

def fetch_emails(service):
    print("Searching Gmail for job-related emails...")
    query = "application OR interview OR hiring OR recruiter OR resume OR rejected OR shortlisted OR \"thank you for applying\" OR \"your application\" OR assessment OR screening OR offer"

    scanned_ids = load_scanned_ids()
    is_first_run = len(scanned_ids) == 0

    if is_first_run:
        print("First run — fetching all emails from the beginning...")
    else:
        print(f"Incremental run — skipping {len(scanned_ids)} already scanned emails...")

    emails = []
    page_token = None

    while len(emails) < MAX_EMAILS:
        batch_size = min(500, MAX_EMAILS - len(emails))
        kwargs = {
            'userId': 'me',
            'q': query,
            'maxResults': batch_size
        }
        if page_token:
            kwargs['pageToken'] = page_token

        results = service.users().messages().list(**kwargs).execute()
        messages = results.get('messages', [])
        if not messages:
            break

        new_messages = [m for m in messages if m['id'] not in scanned_ids]

        if not new_messages and not is_first_run:
            print("  No new emails found in this page, stopping...")
            break

        print(f"  Found {len(new_messages)} new emails in this page...")

        for msg in new_messages:
            try:
                detail = service.users().messages().get(userId='me', id=msg['id'], format='full').execute()
                headers = {h['name']: h['value'] for h in detail['payload']['headers']}
                subject = headers.get('Subject', '')
                sender = headers.get('From', '')
                date = headers.get('Date', '')

                body = ''
                payload = detail['payload']
                if 'parts' in payload:
                    for part in payload['parts']:
                        if part['mimeType'] == 'text/plain' and 'data' in part.get('body', {}):
                            body = base64.urlsafe_b64decode(part['body']['data']).decode('utf-8', errors='ignore')
                            break
                elif 'body' in payload and 'data' in payload['body']:
                    body = base64.urlsafe_b64decode(payload['body']['data']).decode('utf-8', errors='ignore')

                emails.append({
                    'id': msg['id'],
                    'subject': subject,
                    'from': sender,
                    'date': date,
                    'body': body[:1500]
                })
                print(f"  Fetched {len(emails)} new emails so far...", end='\r')
            except Exception as e:
                print(f"  Skipped one email: {e}")
                continue

        page_token = results.get('nextPageToken')
        if not page_token:
            break

    print(f"\nFetched {len(emails)} new emails.")
    return emails

def analyze_emails(emails):
    client = genai.Client(api_key=GEMINI_API_KEY)
    all_jobs = []

    batch_size = 10
    batches = [emails[i:i+batch_size] for i in range(0, len(emails), batch_size)]
    print(f"Analyzing emails in {len(batches)} batches...")

    for i, batch in enumerate(batches):
        print(f"  Analyzing batch {i+1}/{len(batches)}...")
        emails_text = ""
        for j, e in enumerate(batch):
            emails_text += f"\n[{j+1}] Subject: {e['subject']}\nFrom: {e['from']}\nDate: {e['date']}\nBody: {e['body']}\n"

        prompt = f"""Analyze these emails and extract job application data.

For each email that is job-related (application confirmation, interview invite, offer, rejection, recruiter outreach, screening, assessment), extract:
- company: company name
- role: job title applied for
- date: in YYYY-MM-DD format
- found_on: where the job was discovered — LinkedIn / Indeed / Glassdoor / ZipRecruiter / Company Site / Recruiter / Job Alert / Unknown
- applied_via: how they applied — LinkedIn Easy Apply / Workday / Greenhouse / Lever / iCIMS / Company Website / Indeed / Unknown. If the email is from workday.com or mentions Workday anywhere, always use Workday.
- email_type: one of Application Confirmation / Should Apply / Interview Invite / Screening Call / Online Assessment / Offer Letter / Rejection / Unknown
- stage: current stage — one of Applied / Phone Screen / Online Assessment / Interview / Offer / Rejected / Unknown
- description: 1-sentence summary of the role or email
- next_steps: what action is needed e.g. "Follow up", "Prepare for interview", "Complete assessment", "Negotiate offer", "None"
- ai_remarks: a smart, helpful observation about this application. Examples: "No response in over 2 weeks — consider following up", "Rejection received — role no longer active", "Interview invite — act quickly", "Recruiter reached out directly — high priority", "Applied via Easy Apply — lower response rate", "Assessment pending — complete ASAP", "Duplicate application detected", "Offer received — review and negotiate". Keep it short, 1 sentence, actionable.
- suggested_apply: true if a recruiter or job alert is suggesting a role not yet applied to, else false

Skip emails clearly unrelated to jobs.

Return ONLY a valid JSON array, no markdown, no explanation:
[{{"company":"...","role":"...","date":"...","found_on":"...","applied_via":"...","email_type":"...","stage":"...","description":"...","next_steps":"...","ai_remarks":"...","suggested_apply":false}}]

Emails:
{emails_text}"""

        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=prompt
            )
            text = response.text.strip()
            text = text.replace('```json', '').replace('```', '').strip()
            start = text.find('[')
            end = text.rfind(']')
            if start != -1 and end != -1:
                jobs = json.loads(text[start:end+1])
                all_jobs.extend(jobs)
        except Exception as e:
            print(f"  Error in batch {i+1}: {e}")
            continue

    print(f"Extracted {len(all_jobs)} job entries.")
    return all_jobs

def save_to_excel(jobs):
    wb = Workbook()

    # Sheet 1: Job Applications
    ws1 = wb.active
    ws1.title = "Job Applications"

    headers = ["Company", "Role", "Stage", "Date", "Found On", "Applied Via", "Email Type", "Description", "Next Steps", "AI Remarks", "My Notes"]
    header_fill = PatternFill("solid", fgColor="2D2D2D")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    stage_colors = {
        "Applied":           "B5D4F4",  # light blue
        "Interview":         "C0DD97",  # green
        "Online Assessment": "FAC775",  # orange
        "Phone Screen":      "FAF075",  # yellow
        "Offer":             "CECBF6",  # purple
        "Rejected":          "F7C1C1",  # red
        "Unknown":           "D3D1C7"   # gray
    }

    # light yellow background for AI Remarks column
    remarks_fill = PatternFill("solid", fgColor="FFFDE7")
    # light green background for My Notes column
    notes_fill = PatternFill("solid", fgColor="F1F8E9")

    applied_jobs = [j for j in jobs if not j.get('suggested_apply')]
    dates = [j.get('date', '') for j in applied_jobs if j.get('date', '')]
    start_date = min(dates) if dates else "unknown"
    end_date = max(dates) if dates else "unknown"

    for row, job in enumerate(applied_jobs, 2):
        ws1.cell(row=row, column=1, value=job.get('company', '')).font = Font(name="Arial")
        ws1.cell(row=row, column=2, value=job.get('role', '')).font = Font(name="Arial")

        stage = job.get('stage', 'Unknown')
        stage_cell = ws1.cell(row=row, column=3, value=stage)
        stage_cell.fill = PatternFill("solid", fgColor=stage_colors.get(stage, "D3D1C7"))
        stage_cell.font = Font(name="Arial")

        ws1.cell(row=row, column=4, value=job.get('date', '')).font = Font(name="Arial")
        ws1.cell(row=row, column=5, value=job.get('found_on', '')).font = Font(name="Arial")
        ws1.cell(row=row, column=6, value=job.get('applied_via', '')).font = Font(name="Arial")
        ws1.cell(row=row, column=7, value=job.get('email_type', '')).font = Font(name="Arial")
        ws1.cell(row=row, column=8, value=job.get('description', '')).font = Font(name="Arial")
        ws1.cell(row=row, column=9, value=job.get('next_steps', '')).font = Font(name="Arial")

        # AI Remarks — light yellow
        ai_cell = ws1.cell(row=row, column=10, value=job.get('ai_remarks', ''))
        ai_cell.fill = remarks_fill
        ai_cell.font = Font(name="Arial", italic=True, color="7D6608")

        # My Notes — light green, blank for manual entry
        notes_cell = ws1.cell(row=row, column=11, value="")
        notes_cell.fill = notes_fill
        notes_cell.font = Font(name="Arial")

    col_widths = [20, 25, 16, 12, 15, 18, 22, 40, 25, 35, 30]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = width

    # Sheet 2: To Apply
    ws2 = wb.create_sheet("To Apply")
    headers2 = ["Company", "Role", "Found On", "Description", "AI Remarks", "My Notes"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    suggested = [j for j in jobs if j.get('suggested_apply')]
    for row, job in enumerate(suggested, 2):
        ws2.cell(row=row, column=1, value=job.get('company', '')).font = Font(name="Arial")
        ws2.cell(row=row, column=2, value=job.get('role', '')).font = Font(name="Arial")
        ws2.cell(row=row, column=3, value=job.get('found_on', '')).font = Font(name="Arial")
        ws2.cell(row=row, column=4, value=job.get('description', '')).font = Font(name="Arial")

        ai_cell = ws2.cell(row=row, column=5, value=job.get('ai_remarks', ''))
        ai_cell.fill = remarks_fill
        ai_cell.font = Font(name="Arial", italic=True, color="7D6608")

        notes_cell = ws2.cell(row=row, column=6, value="")
        notes_cell.fill = notes_fill
        notes_cell.font = Font(name="Arial")

    for col, width in enumerate([20, 25, 15, 40, 35, 30], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = width

    filename = f"job_tracker_{start_date}_to_{end_date}.xlsx"
    wb.save(filename)
    print(f"\nSaved to {filename}")
    print(f"  - {len(applied_jobs)} applications in 'Job Applications' sheet")
    print(f"  - {len(suggested)} suggestions in 'To Apply' sheet")
    print(f"  - Date range: {start_date} to {end_date}")
    return filename

def main():
    print("=== Job Application Tracker ===\n")
    service = authenticate_gmail()
    emails = fetch_emails(service)
    if not emails:
        print("No new emails found since last run.")
        return
    jobs = analyze_emails(emails)
    if not jobs:
        print("No job data could be extracted.")
        return
    save_to_excel(jobs)
    new_ids = {e['id'] for e in emails}
    save_scanned_ids(new_ids)
    print(f"Saved {len(new_ids)} email IDs to scanned_ids.json")
    print("\nDone!")

if __name__ == '__main__':
    main()